#!/usr/bin/env python3

#======================================================
# Source Modules
#======================================================
from collections import OrderedDict
from class_terraform import terraform_cloud
from easy_functions import countKeys, findKeys, findVars
from easy_functions import create_selector, create_tf_file
from easy_functions import easyDict_append, easyDict_append_subtype
from easy_functions import process_kwargs, query_module_type, query_switch_model
from easy_functions import required_args_add, required_args_remove
from easy_functions import sensitive_var_site_group, stdout_log, validate_args
from easy_functions import variablesFromAPI, vlan_list_full
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import ast
import jinja2
import json
import os
import pkg_resources
import re
import validating

aci_template_path = pkg_resources.resource_filename('classes', 'templates/')

#======================================================
# Exception Classes
#======================================================
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Function.
#=====================================================================================
class access(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - Global Policies - AAEP Profiles
    #======================================================
    def aep_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.global.attachableAccessEntityProfile']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        domain_list = ['physical_domains', 'l3_domains', 'vmm_domains']
        for i in domain_list:
            if not templateVars[f'{i}'] == None:
                if ',' in templateVars[f'{i}']:
                    templateVars[f'{i}'] = templateVars[f'{i}'].split(',')

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'aaep_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - APIC Inband Configuration
    #======================================================
    def apic_inb(self, **kwargs):
        # Dicts for required and optional args
        required_args = {
            'site_group': '',
            'name': '',
            'node_id': '',
            'pod_id': '',
            'Inband_EPG': ''
        }
        optional_args = {
            'Inband_IPv4': '',
            'Inband_GWv4': '',
            'Inband_IPv6': '',
            'Inband_GWv6': '',
        }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Configure the Generic Template Variables
        templateVars['Device_Type'] = 'apic'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Assign the APIC Inband Management IP's
        eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

    #======================================================
    # Function - Interface Policies - CDP
    #======================================================
    def cdp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.cdpInterface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'cdp_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Policies - Fibre Channel
    #======================================================
    def fibre_channel(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.fibreChannelInterface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'fibre_channel_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Profiles
    #======================================================
    def intf_profile(self, **kwargs):
        # Dicts for required and optional args
        required_args = {
            'site_group': '',
            'Switch_Role': '',
            'name': '',
            'Dest_Folder': ''
        }
        optional_args = {'description': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.name_rule('name', templateVars['name'])
            validating.name_rule('Dest_Folder', templateVars['Dest_Folder'])
            validating.values('Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            if not templateVars['description'] == None:
                validating.description('description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        if templateVars['Switch_Role'] == 'leaf':
            # Define the Template Source
            template_file = "leaf_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['name'])
        elif templateVars['Switch_Role'] == 'spine':
            # Define the Template Source
            template_file = "spine_interface_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_Interface_Profile.tf' % (templateVars['name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'

    #======================================================
    # Function - Policy Groups - Interface Policies
    # Shared Policies with Access and Bundle Poicies Groups
    #======================================================
    def interface_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policyGroups.interfacePolicies']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Selectors
    #======================================================
    def intf_selector(self, wb, ws, row_num, wr_file, **kwargs):
        if not kwargs.get('Policy_Group') == None:
            # Dicts for required and optional args
            required_args = {
                'site_group': '',
                'Site_name': '',
                'Switch_Role': '',
                'pod_id': '',
                'node_id': '',
                'Interface_Profile': '',
                'Interface_Selector': '',
                'Port': '',
                'Policy_Group': '',
                'Port_Type': ''
            }
            optional_args = {
                'LACP_Policy': '',
                'Bundle_ID': '',
                'description': '',
                'Switchport_Mode': '',
                'Access_or_Native': '',
                'Trunk_Allowed_VLANs': ''
            }

            kwargs['Switch_Site'] = kwargs.get('site_group')
            if not kwargs.get('Port_Type') == None:
                if re.search('(port-channel|vpc)', kwargs.get('Port_Type')):

                    temp_descr = kwargs['description']
                    # Open the Access Worksheet and Find the Policy Group
                    ws_pg = wb['Access']
                    rows = ws_pg.max_row
                    row_bundle = ''
                    func = 'pg_access'
                    count = countKeys(ws_pg, func)
                    var_dict = findVars(ws_pg, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('name') == kwargs.get('Policy_Group'):
                            row_bundle = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Open the Network Policies Worksheet to get the Interface_Policy
                    ws_net = kwargs['wb']['Network Policies']
                    rows = ws_net.max_row

                    # Get the Interface Policies from the Network Policies Tab
                    func = 'intf_polgrp'
                    count = countKeys(ws_net, func)
                    row_pg = ''
                    var_dict = findVars(ws_net, func, rows, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Policy_name') == kwargs.get('Interface_Policy'):
                            row_pg = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Validate inputs, return dict of template vars

                    if kwargs.get('Port_Type') == 'vpc':
                        kwargs['Lag_Type'] = 'node'
                        ws_vpc = wb['Inventory']
                        for row in ws_vpc.rows:
                            if row[0].value == 'vpc_pair' and int(row[1].value) == int(kwargs.get('Switch_Site')) and str(row[5].value) == str(kwargs.get('node_id')):
                                kwargs['VPC_name'] = row[2].value
                                kwargs['name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))

                            elif row[0].value == 'vpc_pair' and str(row[1].value) == str(kwargs.get('Switch_Site')) and str(row[6].value) == str(kwargs.get('node_id')):
                                kwargs['VPC_name'] = row[2].value
                                kwargs['name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))
                    elif kwargs.get('Port_Type') == 'port-channel':
                        kwargs['Lag_Type'] = 'link'
                        kwargs['name'] = '%s_pc%s' % (kwargs.get('Interface_Profile'), kwargs.get('Bundle_ID'))

                    kwargs['description'] = temp_descr
                    # Create the Bundle Policy Group
                    kwargs['site_group'] = kwargs.get('Switch_Site')
                    lib_aci_ref = 'Access_Policies'
                    class_init = '%s(ws)' % (lib_aci_ref)
                    func = 'pg_bundle'
                    eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, func))

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)
            # leafx = name
            xa = templateVars['Port'].split('/')
            xcount = len(xa)
            templateVars['Module_From'] = xa[0]
            templateVars['Module_To'] = xa[0]
            templateVars['Port_From'] = xa[1]
            templateVars['Port_To'] = xa[1]
            templateVars['Selector_Type'] = 'range'

            if templateVars['Switch_Role'] == 'leaf':
                templateVars['Policy_Group'] = kwargs.get('name')
                # Define the Template Source
                template_file = "leaf_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if xcount == 3:
                    templateVars['SubPort_From'] = xa[2]
                    templateVars['SubPort_To'] = xa[2]
                    template_file = "leaf_portblock_sub.jinja2"
                else:
                    template_file = "leaf_portblock.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

            elif templateVars['Switch_Role'] == 'spine':
                # Define the Template Source
                template_file = "spine_portselect.jinja2"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

    #======================================================
    # Function - Interface Policies - L2 Interfaces
    #======================================================
    def l2_interface(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.L2Interface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'l2_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Domain - Layer 3
    #======================================================
    def l3_domain(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.domains.Layer3']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'layer3_domains'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Leaf Policy Group
    #======================================================
    def leaf_pg(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.switches.leafPolicyGroup']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'leaf_policy_groups'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Policies - Link Level (Speed)
    #======================================================
    def link_level(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.linkLevel']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'link_level_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']
    #======================================================
    # Function - Interface Policies - LLDP
    #======================================================
    def lldp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.lldpInterface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'lldp_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Policies - Mis-Cabling Protocol
    #======================================================
    def mcp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.mcpInterface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'mcp_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Static Management IPs
    #======================================================
    def mgmt_static(self, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'name': '',
                         'node_id': '',
                         'pod_id': '',
                         'Device_Type': '',
                         'Type': '',
                         'EPG': ''}
        optional_args = {'IPv4': '',
                         'GWv4': '',
                         'IPv6': '',
                         'GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.hostname('name', templateVars['name'])
            validating.name_rule('EPG', templateVars['EPG'])
            validating.number_check('pod_id', templateVars['pod_id'], 1, 15)
            if templateVars['Device_Type'] == 'apic':
                validating.number_check('node_id', templateVars['node_id'], 1, 7)
            else:
                validating.number_check('node_id', templateVars['node_id'], 101, 4001)
            if not templateVars['IPv4'] == None:
                validating.mgmt_network('IPv4', templateVars['IPv4'], 'GWv4', templateVars['GWv4'])
            if not templateVars['IPv6'] == None:
                validating.mgmt_network('IPv6', templateVars['IPv6'], 'GWv6', templateVars['GWv6'])
            else:
                templateVars['IPv6'] = '::'
                templateVars['GWv6'] = '::'
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "static_node_mgmt_address.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_%s_EPG_%s_Static_Address.tf' % (templateVars['name'], templateVars['Type'], templateVars['EPG'])
        dest_dir = 'Tenant_mgmt'
        
    #======================================================
    # Function - Policy Group - Access
    #======================================================
    def pg_access(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policyGroups.leafAccessPort']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        for item in kwargs['easyDict']['access']['interface_policies']:
            for key, value in item.items():
                if key == kwargs['site_group']:
                    for i in value:
                        if i['interface_policy'] == templateVars['interface_policy']:
                            templateVars.update(i)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'leaf_port_group_access'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Policy Group - VPC/Port-Channel
    #======================================================
    def pg_bundle(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policyGroups.leafBundle']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        for item in kwargs['easyDict']['access']['interface_policies']:
            for key, value in item.items():
                if key == kwargs['site_group']:
                    for i in value:
                        if i['interface_policy'] == templateVars['interface_policy']:
                            templateVars.update(i)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'leaf_port_group_bundle'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Policy Group - Breakout
    #======================================================
    def pg_breakout(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policyGroups.leafBreakOut']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'leaf_port_group_breakout'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Policy Group - Spine
    #======================================================
    def pg_spine(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policyGroups.spineAccessPort']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'spine_port_group_access'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Domains - Physical
    #======================================================
    def phys_domain(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.domains.Physical']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'physical_domains'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Policies - Port Channel
    #======================================================
    def port_channel(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.PortChannel']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'port_channel_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Port Conversion
    #======================================================
    def port_cnvt(self, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'name': '',
                         'node_id': '',
                         'Port': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.hostname('name', templateVars['name'])
            validating.number_check('node_id', templateVars['node_id'], 101, 4001)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Create Port name Var
        zz = templateVars['Port'].split('/')
        templateVars['Port_name'] = '%s_%s' % (zz[0], zz[1])

        # Define the Template Source
        template_file = "downlink.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'Downlink_Convert_%s.tf' % (templateVars['Port_name'])
        dest_dir = 'Access/%s' % (templateVars['name'])
        
    #======================================================
    # Function - Interface Policies - Port Security
    #======================================================
    def port_security(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.portSecurity']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'port_security_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Spine Policy Group
    #======================================================
    def spine_pg(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.switches.spinePolicyGroup']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'spine_policy_groups'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Interface Policies - Spanning Tree
    #======================================================
    def stp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.policies.spanningTreeInterface']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'spanning_tree_interface_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Switch Inventory
    #======================================================
    def switch(self, **kwargs):
        # Initialize the Class
        lib_aci_ref = 'Access_Policies'
        class_init = '%s(ws)' % (lib_aci_ref)

        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Serial': '',
                         'node_id': '',
                         'name': '',
                         'Profiles': '',
                         'Node_Type': '',
                         'pod_id': '',
                         'Switch_Role': '',
                         'Switch_Type': '',
                         'Is_Virtual': '',
                         'Tier-2': '',
                         'Inband_EPG': '',
                         'OOB_EPG': ''}
        optional_args = {'Policy_Group': '',
                         'Remote_ID': '',
                         'Fabric_ID': '',
                         'MG_name': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',
                         'OOB_IPv4': '',
                         'OOB_GWv4': '',
                         'OOB_IPv6': '',
                         'OOB_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Use Switch_Type to Determine the Number of ports on the switch
        modules,port_count = query_switch_model(row_num, kwargs['Switch_Type'])

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.hostname('name', templateVars['name'])
            validating.modules(row_num, templateVars['name'], templateVars['Switch_Role'], modules)
            if not templateVars['MG_name'] == None:
                validating.name_rule('MG_name', templateVars['MG_name'])
            validating.name_rule('Inband_EPG', templateVars['Inband_EPG'])
            validating.name_rule('OOB_EPG', templateVars['OOB_EPG'])
            validating.number_check('node_id', templateVars['node_id'], 101, 4001)
            validating.number_check('pod_id', templateVars['pod_id'], 1, 12)
            validating.number_check('Fabric_ID', templateVars['Fabric_ID'], 1, 12)
            validating.values('Profiles', templateVars['Profiles'], ['no', 'yes'])
            validating.values('Node_Type', templateVars['Node_Type'], ['remote-leaf-wan', 'unspecified'])
            validating.values('Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values('Is_Virtual', templateVars['Is_Virtual'], ['no', 'yes'])
            validating.values('Tier-2', templateVars['Tier-2'], ['no', 'yes'])
            if templateVars['Profiles'] == 'yes':
                validating.name_rule('Policy_Group', templateVars['Policy_Group'])
            if not templateVars['Remote_ID'] == None:
                validating.number_check('Remote_ID', templateVars['Remote_ID'], 1, 255)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Configure the Generic Template Variables for Node Management Inband IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'in_band'
        templateVars['EPG'] = templateVars['Inband_EPG']
        templateVars['IPv4'] = templateVars['Inband_IPv4']
        templateVars['GWv4'] = templateVars['Inband_GWv4']
        templateVars['IPv6'] = templateVars['Inband_IPv6']
        templateVars['GWv6'] = templateVars['Inband_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Inband Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        # Configure the Generic Template Variables for Node Management Out-of-Band IP's
        templateVars['Device_Type'] = 'switch'
        templateVars['Type'] = 'out_of_band'
        templateVars['EPG'] = templateVars['OOB_EPG']
        templateVars['IPv4'] = templateVars['OOB_IPv4']
        templateVars['GWv4'] = templateVars['OOB_GWv4']
        templateVars['IPv6'] = templateVars['OOB_IPv6']
        templateVars['GWv6'] = templateVars['OOB_GWv6']

        if re.search('\d', templateVars['IPv4']) or re.search('\d', templateVars['IPv6']):
            # Assign the Switch Out-of-Band Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_static'))

        if not templateVars['MG_name'] == None:
            # Define the Template Source
            template_file = "maint_group_nodeblk.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'Maintenance_Group_%s.tf' % (templateVars['MG_name'])
            dest_dir = 'Admin'
            

        Site_ID = 'Site_ID_%s' % (templateVars['site_group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create kwargs for Site Variables
        templateVars['Site_ID'] = site_dict.get('Site_ID')
        templateVars['Site_name'] = site_dict.get('Site_name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        templateVars['APIC_Version'] = site_dict.get('APIC_Version')
        templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
        templateVars['Terraform_EQ'] = site_dict.get('Terraform_EQ')
        templateVars['Terraform_Version'] = site_dict.get('Terraform_Version')
        templateVars['Provider_EQ'] = site_dict.get('Provider_EQ')
        templateVars['Provider_Version'] = site_dict.get('Provider_Version')
        templateVars['Run_Location'] = site_dict.get('Run_Location')
        templateVars['State_Location'] = site_dict.get('State_Location')
        templateVars['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
        templateVars['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
        templateVars['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
        templateVars['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

        self.templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        excel_wkbook = '%s_intf_selectors.xlsx' % (templateVars['Site_name'])

        wb_sw = load_workbook(excel_wkbook)

        # Check if there is a Worksheet for the Switch Already
        if not templateVars['name'] in wb_sw.sheetnames:
            ws_sw = wb_sw.create_sheet(title = templateVars['name'])
            ws_sw = wb_sw[templateVars['name']]
            ws_sw.column_dimensions['A'].width = 15
            ws_sw.column_dimensions['B'].width = 10
            ws_sw.column_dimensions['c'].width = 10
            ws_sw.column_dimensions['d'].width = 20
            ws_sw.column_dimensions['E'].width = 20
            ws_sw.column_dimensions['F'].width = 10
            ws_sw.column_dimensions['G'].width = 20
            ws_sw.column_dimensions['H'].width = 20
            ws_sw.column_dimensions['I'].width = 20
            ws_sw.column_dimensions['J'].width = 15
            ws_sw.column_dimensions['K'].width = 30
            ws_sw.column_dimensions['L'].width = 20
            ws_sw.column_dimensions['M'].width = 20
            ws_sw.column_dimensions['N'].width = 30
            dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
            dv2 = DataValidation(type="list", formula1='"access,breakout,port-channel,vpc"', allow_blank=True)
            ws_sw.add_data_validation(dv1)
            ws_sw.add_data_validation(dv2)
            ws_header = '%s Interface Selectors' % (templateVars['name'])
            data = [ws_header]
            ws_sw.append(data)
            ws_sw.merge_cells('A1:N1')
            for cell in ws_sw['1:1']:
                cell.style = 'Heading 1'
            data = ['','Notes: Breakout Policy Group names are 2x100g_pg, 4x10g_pg, 4x25g_pg, 4x100g_pg, 8x50g_pg.']
            ws_sw.append(data)
            ws_sw.merge_cells('B2:N2')
            for cell in ws_sw['2:2']:
                cell.style = 'Heading 2'
            data = ['Type','pod_id','node_id','Interface_Profile','Interface_Selector','Port','Policy_Group','Port_Type','LACP_Policy','Bundle_ID','description','Switchport_Mode','Access_or_Native','Trunk_Allowed_VLANs']
            ws_sw.append(data)
            for cell in ws_sw['3:3']:
                cell.style = 'Heading 3'

            ws_sw_row_count = 4
            templateVars['dv1'] = dv1
            templateVars['dv2'] = dv2
            templateVars['port_count'] = port_count
            sw_type = str(templateVars['Switch_Type'])
            sw_name = str(templateVars['name'])
            if re.search('^(93[0-9][0-9])', sw_type):
                for module in range(1, 2):
                    templateVars['module'] = module
                    ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
            if re.search('^(9396|95[0-1][4-8])', sw_type):
                row_count = 1
                for row in kwargs['ws'].rows:
                    if re.search('9396', sw_type):
                        start, end = 2, 2
                    else:
                        start, end = 1, int(modules)
                    if str(row[0].value) == sw_type and str(row[2].value) == sw_name:
                        for module in range(start, end + 2):
                            templateVars['module'] = module
                            module_type = row[module + 2].value
                            if module_type == None:
                                module_type = 'none'
                            elif re.search('(X97|M(4|6|12)P)', module_type):
                                templateVars['port_count'] = query_module_type(row_count, module_type)
                                ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
                        row_count += 1
                        break
            wb_sw.save(excel_wkbook)
        else:
            ws_sw = wb_sw[templateVars['name']]

        # Define the Template Source
        template_file = "inventory.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['name'])
        dest_dir = '%s' % (templateVars['name'])
        

        if templateVars['Profiles'] == 'yes':
            templateVars['description'] = None
            templateVars['Dest_Folder'] = templateVars['name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'intf_profile'))

            templateVars['Selector_name'] = templateVars['name']
            templateVars['Association_Type'] = 'range'
            templateVars['Nodeblk_name'] = 'blk%s-%s' % (templateVars['node_id'], templateVars['node_id'])
            templateVars['node_id_From'] = templateVars['node_id']
            templateVars['node_id_To'] = templateVars['node_id']
            templateVars['Interface_Profile'] = templateVars['name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'sw_profile'))

            sw_intf_profile = './ACI/%s/%s/%s_Interface_Profile.tf' % (templateVars['Site_name'], templateVars['name'], templateVars['name'])
            wr_file = open(sw_intf_profile, 'a+')
            lib_aci_ref = 'Access_Policies'
            rows_sw = ws_sw.max_row
            func_regex = re.compile('^intf_selector$')
            func_list = findKeys(ws_sw, func_regex)
            class_init = '%s(ws_sw)' % (lib_aci_ref)
            stdout_log(ws_sw, None, 'begin')
            for func in func_list:
                count = countKeys(ws_sw, func)
                var_dict = findVars(ws_sw, func, rows_sw, count)
                for pos in var_dict:
                    row_num = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    for x in list(var_dict[pos].keys()):
                        if var_dict[pos][x] == '':
                            del var_dict[pos][x]
                    stdout_log(ws_sw, row_num, 'begin')
                    var_dict[pos]['site_group'] = templateVars['Site_ID']
                    var_dict[pos]['Switch_Role'] = templateVars['Switch_Role']
                    var_dict[pos]['Site_name'] = templateVars['Site_name']
                    eval("%s.%s(wb, ws_sw, row_num, wr_file, **var_dict[pos])" % (class_init, func))
            wr_file.close()
            ws_wr = wb_sw.get_sheet_names()
            for sheetname in ws_wr:
                if sheetname in ['Sites']:
                    sheetToDelete = wb_sw.get_sheet_by_name(sheetname)
                    wb_sw.remove_sheet(sheetToDelete)
                    wb_sw.save(excel_wkbook)
            wb_sw.close()

        if re.search('Grp_[A-F]', templateVars['site_group']):
            # print(f"\n-----------------------------------------------------------------------------\n")
            # print(f"   Error on Worksheet {ws.title}, Row {row_num} site_group, value {templateVars['site_group']}.")
            # print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            # print(f"\n-----------------------------------------------------------------------------\n")
            exit()
        elif re.search(r'\d+', templateVars['site_group']):
            Site_ID = 'Site_ID_%s' % (templateVars['site_group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Set Destination Directory
            dest_dir = '%s' % (templateVars['name'])

            # Create kwargs for Site Variables
            kwargs['Site_ID'] = site_dict.get('Site_ID')
            kwargs['Site_name'] = site_dict.get('Site_name')
            kwargs['APIC_URL'] = site_dict.get('APIC_URL')
            kwargs['APIC_Version'] = site_dict.get('APIC_Version')
            kwargs['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
            kwargs['Terraform_EQ'] = site_dict.get('Terraform_EQ')
            kwargs['Terraform_Version'] = site_dict.get('Terraform_Version')
            kwargs['Provider_EQ'] = site_dict.get('Provider_EQ')
            kwargs['Provider_Version'] = site_dict.get('Provider_Version')
            kwargs['Run_Location'] = site_dict.get('Run_Location')
            kwargs['State_Location'] = site_dict.get('State_Location')
            kwargs['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
            kwargs['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
            kwargs['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
            kwargs['Terraform_Agent_Pool_ID'] = site_dict.get('Terraform_Agent_Pool_ID')

            # Dicts for required and optional args
            required_args = {'Site_ID': '',
                                'Site_name': '',
                                'APIC_URL': '',
                                'APIC_Version': '',
                                'APIC_Auth_Type': '',
                                'Terraform_EQ': '',
                                'Terraform_Version': '',
                                'Provider_EQ': '',
                                'Provider_Version': '',
                                'Run_Location': '',
                                'State_Location': ''}
            optional_args = {'Terraform_Cloud_Org': '',
                                'Workspace_Prefix': '',
                                'VCS_Base_Repo': '',
                                'Terraform_Agent_Pool_ID': ''}

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)

            # If the State_Location is Terraform_Cloud Configure Workspaces in the Cloud
            if templateVars['State_Location'] == 'Terraform_Cloud':
                # Initialize the Class
                lib_tf_ref = 'lib_terraform.Terraform_Cloud'
                class_init = '%s()' % (lib_tf_ref)

                # Get terraform_cloud_token
                kwargs['terraform_cloud_token'] = eval("%s.%s()" % (class_init, 'terraform_token'))

                # Get terraform_cloud_token
                kwargs['terraform_oath_token'] = eval("%s.%s(**kwargs)" % (class_init, 'oath_token'))

                # Get workspace_ids
                workspace_dict = {}
                workspace_dict = terraform_cloud.tfcWorkspace(class_init, dest_dir, workspace_dict, **kwargs)

            # If the Run_Location is Terraform_Cloud Configure Variables in the Cloud
            if templateVars['Run_Location'] == 'Terraform_Cloud':
                # Set Variable List
                if templateVars['APIC_Auth_Type'] == 'user_pass':
                    var_list = ['aciUrl', 'aciUser', 'aciPass']
                else:
                    var_list = ['aciUrl', 'aciCertname', 'aciPrivateKey']

                # Get var_ids
                tf_var_dict = {}
                folder_id = 'Site_ID_%s_%s' % (templateVars['Site_ID'], dest_dir)
                kwargs['workspace_id'] = workspace_dict[folder_id]
                kwargs['description'] = ''
                for var in var_list:
                    tf_var_dict = terraform_cloud.tfcVariables(class_init, dest_dir, var, tf_var_dict, **kwargs)

        # else:
        #     print(f"\n-----------------------------------------------------------------------------\n")
        #     print(f"   Error on Worksheet {ws.title}, Row {row_num} site_group, value {templateVars['site_group']}.")
        #     print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        #     print(f"\n-----------------------------------------------------------------------------\n")
        #     exit()

        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

        # Add the Default Files to the Tenant Directory
        file_list = ['.gitignore_.gitignore', 'main.jinja2_main.tf', 'variables.jinja2_variables.tf']
        for file in file_list:
            x = file.split('_')
            template_file = x[0]
            dest_file = x[1]
            template = self.templateEnv.get_template(template_file)
            create_tf_file('w', dest_dir, dest_file, template, **templateVars)

    #======================================================
    # Function - Switch Profiles
    #======================================================
    def sw_profile(self, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'Switch_Role': '',
                         'name': '',
                         'Selector_name': '',
                         'Association_Type': '',
                         'Nodeblk_name': '',
                         'node_id_From': '',
                         'node_id_To': '',
                         'Policy_Group': '',
                         'Interface_Profile': '',
                         'Dest_Folder': ''}
        optional_args = {'description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.name_rule('name', templateVars['name'])
            validating.name_rule('Selector_name', templateVars['Selector_name'])
            validating.name_rule('Nodeblk_name', templateVars['Nodeblk_name'])
            validating.name_rule('Policy_Group', templateVars['Policy_Group'])
            validating.name_rule('Interface_Profile', templateVars['Interface_Profile'])
            validating.name_rule('Dest_Folder', templateVars['Dest_Folder'])
            validating.number_check('node_id_From', templateVars['node_id_From'], 101, 4001)
            validating.number_check('node_id_To', templateVars['node_id_To'], 101, 4001)
            validating.values('Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values('Association_Type', templateVars['Association_Type'], ['ALL', 'range', 'ALL_IN_POD'])
            if not templateVars['description'] == None:
                validating.description('description', templateVars['description'])
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Define the Template Source
        if templateVars['Switch_Role'] == 'leaf':
            template_file = "leaf_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_leaf_profile.tf' % (templateVars['name'])
        else:
            template_file = "spine_profile.jinja2"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_spine_profile.tf' % (templateVars['name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'

    #======================================================
    # Function - VLAN Pools
    #======================================================
    def vlan_pool(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.pools.Vlan']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'vlan_pools'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Virtual Networking - Controllers
    #======================================================
    def vmm_controllers(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.vmm.Controllers']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to Policy
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'virtual_networking'
        templateVars['data_subtype'] = 'controllers'
        templateVars['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Virtual Networking - Credentials
    #======================================================
    def vmm_creds(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.vmm.Credentials']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'vmm_password_{kwargs["password"]}'
        sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'virtual_networking'
        templateVars['data_subtype'] = 'credentials'
        templateVars['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Virtual Networking - Domains
    #======================================================
    def vmm_domain(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.vmm.Domains']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Convert to Lists
        if not templateVars["uplink_names"] == None:
            if ',' in templateVars["uplink_names"]:
                templateVars["uplink_names"] = templateVars["uplink_names"].split(',')
        else:
            templateVars["uplink_names"] = []

        upDating = {
            'controllers':[],
            'credentials':[],
            'enhanced_lag_policy':[],
            'domain':[templateVars],
            'vswitch_policy':[]
        }
        templateVars = upDating
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'virtual_networking'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Virtual Networking - Controllers
    #======================================================
    def vmm_elagp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.vmm.enhancedLag']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'virtual_networking'
        templateVars['data_subtype'] = 'enhanced_lag_policy'
        templateVars['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Virtual Networking - Controllers
    #======================================================
    def vmm_vswitch(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['access.vmm.vswitchPolicy']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to Policy
        templateVars['class_type'] = 'access'
        templateVars['data_type'] = 'virtual_networking'
        templateVars['data_subtype'] = 'vswitch_policy'
        templateVars['policy_name'] = kwargs['domain_name']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - VPC Pair
    #======================================================
    def vpc_pair(self, **kwargs):
        # Dicts for required and optional args
        required_args = {'site_group': '',
                         'VPC_ID': '',
                         'name': '',
                         'Node1_ID': '',
                         'Node2_ID': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Arguments
            validating.site_group('site_group', **kwargs)
            validating.number_check('VPC_ID', templateVars['VPC_ID'], 1, 1000)
            validating.number_check('Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.number_check('Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Define the Template Source
        template_file = "vpc_domain.jinja2"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vpc_domain_%s.tf' % (templateVars['VPC_ID'])
        dest_dir = 'Access'
        
#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Function.
#=====================================================================================
class admin(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - Authentication
    #======================================================
    def auth(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.Authentication']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        if 'console_realm' in kwargs:
            if not kwargs['console_realm'] == 'local':
                jsonData = required_args_add(['console_login_domain'], jsonData)
        else:
            kwargs['console_realm'] == 'local'
        if 'default_realm' in kwargs:
            if not kwargs['default_realm'] == 'local':
                jsonData = required_args_add(['default_login_domain'], jsonData)
        else:
            kwargs['default_realm'] == 'local'
        
        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Reset jsonData
        if not kwargs['console_realm'] == 'local':
            jsonData = required_args_remove(['console_login_domain'], jsonData)
        if not kwargs['default_realm'] == 'local':
            jsonData = required_args_remove(['default_login_domain'], jsonData)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'authentication'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Configuration Backup - Export Policies
    #======================================================
    def export_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.exportPolicy']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        Additions = {
            'configuration_export': [],
        }
        templateVars.update(Additions)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'configuration_backups'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Global Security Settings
    #======================================================
    def mg_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.firmware.Policy']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)
        templateVars['maintenance_groups'] = []

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'firmware'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Configuration Backup  - Remote Host
    #======================================================
    def maint_group(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.firmware.MaintenanceGroups']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Convert to Lists
        if ',' in templateVars["node_list"]:
            templateVars["node_list"] = templateVars["node_list"].split(',')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'firmware'
        templateVars['data_subtype'] = 'maintenance_groups'
        templateVars['policy_name'] = kwargs['maintenance_group_policy']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - RADIUS Authentication
    #======================================================
    def radius(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.Radius']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        if 'server_monitoring' in kwargs:
            if kwargs['server_monitoring'] == 'enabled':
                jsonData = required_args_add(['monitoring_password', 'username'], jsonData)
        else:
            kwargs['server_monitoring'] == 'disabled'
        
        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the secert is in the Environment.  If not Add it.
        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'radius_key_{kwargs["key"]}'
        sensitive_var_site_group(**templateVars)
        if templateVars['server_monitoring'] == 'enabled':
            # Check if the Password is in the Environment.  If not Add it.
            templateVars["Variable"] = f'radius_monitoring_password_{kwargs["monitoring_password"]}'
            sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Reset jsonData
        if not kwargs['server_monitoring'] == 'disabled':
            jsonData = required_args_remove(['monitoring_password', 'username'], jsonData)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'radius'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Configuration Backup  - Remote Host
    #======================================================
    def remote_host(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.remoteHost']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if 'authentication_type' in kwargs:
            if kwargs['authentication_type'] == 'usePassword':
                jsonData = required_args_add(['username'], jsonData)
        else:
            kwargs['authentication_type'] == 'usePassword'
            jsonData = required_args_add(['username'], jsonData)

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        templateVars['jsonData'] = jsonData
        if templateVars['authentication_type'] == 'usePassword':
            # Check if the Password is in the Environment.  If not Add it.
            templateVars["Variable"] = f'remote_password_{kwargs["password"]}'
            sensitive_var_site_group(**templateVars)
        else:
            # Check if the SSH Key/Passphrase is in the Environment.  If not Add it.
            templateVars["Variable"] = 'ssh_key_contents'
            sensitive_var_site_group(**templateVars)
            templateVars["Variable"] = 'ssh_key_passphrase'
            sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Reset jsonData
        if kwargs['authentication_type'] == 'usePassword':
            jsonData = required_args_remove(['username'], jsonData)
        
        # Convert to Lists
        if ',' in templateVars["remote_hosts"]:
            templateVars["remote_hosts"] = templateVars["remote_hosts"].split(',')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'configuration_backups'
        templateVars['data_subtype'] = 'configuration_export'
        templateVars['policy_name'] = kwargs['scheduler_name']
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Global Security Settings
    #======================================================
    def security(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.globalSecurity']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'security'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - TACACS+ Authentication
    #======================================================
    def tacacs(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['admin.Tacacs']['allOf'][1]['properties']
        
        # Check for Variable values that could change required arguments
        if 'server_monitoring' in kwargs:
            if kwargs['server_monitoring'] == 'enabled':
                jsonData = required_args_add(['monitoring_password', 'username'], jsonData)
        else:
            kwargs['server_monitoring'] == 'disabled'
        
        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the secert is in the Environment.  If not Add it.
        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'tacacs_key_{kwargs["key"]}'
        sensitive_var_site_group(**templateVars)
        if templateVars['server_monitoring'] == 'enabled':
            # Check if the Password is in the Environment.  If not Add it.
            templateVars["Variable"] = f'tacacs_monitoring_password_{kwargs["monitoring_password"]}'
            sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Reset jsonData
        if not kwargs['server_monitoring'] == 'disabled':
            jsonData = required_args_remove(['monitoring_password', 'username'], jsonData)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'admin'
        templateVars['data_type'] = 'tacacs'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Function.
#=====================================================================================
class fabric(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - Date and Time Policy
    #======================================================
    def date_time(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.DateandTime']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # If NTP server state is disabled disable master mode as well.
        if templateVars['server_state'] == 'disabled':
            templateVars['master_mode'] = 'disabled'
        
        Additions = {
            'authentication_keys': [],
            'name':'default',
            'ntp_servers': [],
        }
        templateVars.update(Additions)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'date_and_time'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - DNS Profiles
    #======================================================
    def dns_profile(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.dnsProfiles']['allOf'][1]['properties']

        if not kwargs['dns_domains'] == None:
            if ',' in kwargs['dns_domains']:
                kwargs['dns_domains'] = kwargs['dns_domains'].split(',')
            else:
                kwargs['dns_domains'] = [kwargs['dns_domains']]
            if not kwargs['default_domain'] == None:
                if not kwargs['default_domain'] in kwargs['dns_domains']:
                    kwargs['dns_domains'].append(kwargs['default_domain'])

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        Additions = {
            'name':'default',
        }
        templateVars.update(Additions)
        
        # Convert to Lists
        if ',' in templateVars["dns_providers"]:
            templateVars["dns_providers"] = templateVars["dns_providers"].split(',')

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'dns_profiles'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Date and Time Policy - NTP Servers
    #======================================================
    def ntp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.Ntp']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'date_and_time'
        templateVars['data_subtype'] = 'ntp_servers'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Date and Time Policy - NTP Keys
    #======================================================
    def ntp_key(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.NtpKeys']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the NTP Key is in the Environment.  If not Add it.
        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'ntp_key_{kwargs["key_id"]}'
        sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'date_and_time'
        templateVars['data_subtype'] = 'authentication_keys'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Smart CallHome Policy
    #======================================================
    def smart_callhome(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.smartCallHome']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        Additions = {
            'name':'default',
            'smtp_server': [],
            'smart_destinations': [],
        }
        templateVars.update(Additions)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'smart_callhome'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Smart CallHome Policy - Smart Destinations
    #======================================================
    def smart_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.smartDestinations']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'smart_callhome'
        templateVars['data_subtype'] = 'smart_destinations'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Smart CallHome Policy - SMTP Server
    #======================================================
    def smart_smtp_server(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.smartSmtpServer']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if 'secure_smtp' in kwargs:
            if 'true' in kwargs['secure_smtp']:
                jsonData = required_args_add(['username'], jsonData)
        else:
            kwargs['secure_smtp'] == 'false'

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the Smart CallHome SMTP Password is in the Environment and if not add it.
        if 'true' in kwargs['secure_smtp']:
            templateVars['jsonData'] = jsonData
            templateVars["Variable"] = f'smtp_password'
            sensitive_var_site_group(**templateVars)
            templateVars.pop('jsonData')
            templateVars.pop('Variable')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'smart_callhome'
        templateVars['data_subtype'] = 'smtp_server'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - SNMP Policy - Client Groups
    #======================================================
    def snmp_clgrp(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.snmpClientGroups']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Convert to Lists
        if ',' in templateVars["clients"]:
            templateVars["clients"] = templateVars["clients"].split(',')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'snmp_policies'
        templateVars['data_subtype'] = 'snmp_client_groups'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - SNMP Policy - Communities
    #======================================================
    def snmp_community(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.snmpCommunities']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the SNMP Community is in the Environment.  If not Add it.
        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'snmp_community_{kwargs["community_variable"]}'
        sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'snmp_policies'
        templateVars['data_subtype'] = 'snmp_communities'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - SNMP Policy - SNMP Trap Destinations
    #======================================================
    def snmp_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.snmpDestinations']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if 'version' in kwargs:
            if re.fullmatch('(v1|v2c)', kwargs['version']):
                jsonData = required_args_add(['community_variable'], jsonData)
            elif 'v3' in kwargs['version']:
                jsonData = required_args_add(['username', 'v3_security_level'], jsonData)
        else:
            kwargs['version'] = 'v2c'

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Reset Arguments
        if re.fullmatch('(v1|v2c)', kwargs['version']):
            jsonData = required_args_remove(['community_variable'], jsonData)
        elif 'v3' in kwargs['version']:
            jsonData = required_args_remove(['username', 'v3_security_level'], jsonData)
        
        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'snmp_policies'
        templateVars['data_subtype'] = 'snmp_destinations'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - SNMP Policy
    #======================================================
    def snmp_policy(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.snmpPolicy']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        Additions = {
            'name':'default',
            'snmp_client_groups': [],
            'snmp_communities': [],
            'snmp_destinations': [],
            'users': [],
        }
        templateVars.update(Additions)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'snmp_policies'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - SNMP Policy - SNMP Users
    #======================================================
    def snmp_user(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.snmpUsers']['allOf'][1]['properties']

        # Check for Variable values that could change required arguments
        if 'privacy_key' in kwargs:
            if not kwargs['privacy_key'] == 'none':
                jsonData = required_args_add(['privacy_key'], jsonData)
        else:
            kwargs['privacy_key'] = 'none'

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Check if the Authorization and Privacy Keys are in the environment and if not add them.
        templateVars['jsonData'] = jsonData
        templateVars["Variable"] = f'snmp_authorization_key_{kwargs["authorization_key"]}'
        sensitive_var_site_group(**templateVars)
        if not kwargs['privacy_type'] == 'none':
            templateVars["Variable"] = f'snmp_privacy_key_{kwargs["privacy_key"]}'
            sensitive_var_site_group(**templateVars)
        templateVars.pop('jsonData')
        templateVars.pop('Variable')

        # Reset jsonData
        if not kwargs['privacy_key'] == 'none':
            jsonData = required_args_remove(['privacy_key'], jsonData)
        
        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'snmp_policies'
        templateVars['data_subtype'] = 'users'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Syslog Policy
    #======================================================
    def syslog(self, **kwargs):
       # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.Syslog']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        Additions = {
            'name':'default',
            'remote_destinations': []
        }
        templateVars.update(Additions)
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'syslog'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Syslog Policy - Syslog Destinations
    #======================================================
    def syslog_destinations(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['fabric.syslogRemoteDestinations']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to Policy
        templateVars['class_type'] = 'fabric'
        templateVars['data_type'] = 'syslog'
        templateVars['data_subtype'] = 'remote_destinations'
        templateVars['policy_name'] = 'default'
        kwargs['easyDict'] = easyDict_append_subtype(templateVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Function.
#=====================================================================================
class system_settings(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - APIC Connectivity Preference
    #======================================================
    def apic_preference(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['system.apicConnectivityPreference']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'system_settings'
        templateVars['data_type'] = 'apic_connectivity_preference'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - BGP Autonomous System Number
    #======================================================
    def bgp_asn(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['system.bgpASN']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'system_settings'
        templateVars['data_type'] = 'bgp_asn'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - BGP Route Reflectors
    #======================================================
    def bgp_rr(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['system.bgpRouteReflector']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        # Convert to Lists
        if ',' in templateVars["node_list"]:
            templateVars["node_list"] = templateVars["node_list"].split(',')

        # Add Dictionary to easyDict
        templateVars['class_type'] = 'system_settings'
        templateVars['data_type'] = 'bgp_rr'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

    #======================================================
    # Function - Global AES Passphrase Encryption Settings
    #======================================================
    def global_aes(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['system.globalAesEncryptionSettings']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        if kwargs['enable_encryption'] == 'true':
            templateVars["Variable"] = 'aes_passphrase'
            templateVars['jsonData'] = jsonData
            sensitive_var_site_group(**templateVars)
            templateVars.pop('jsonData')
            templateVars.pop('Variable')
        
        # Add Dictionary to easyDict
        templateVars['class_type'] = 'system_settings'
        templateVars['data_type'] = 'global_aes_encryption_settings'
        kwargs['easyDict'] = easyDict_append(templateVars, **kwargs)
        return kwargs['easyDict']

#=====================================================================================
# Please Refer to the "Notes" in the relevant column headers in the input Spreadhseet
# for detailed information on the Arguments used by this Function.
#=====================================================================================
class site_policies(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - Site Settings
    #======================================================
    def site_id(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['site.Identifiers']['allOf'][1]['properties']

        try:
            # Validate User Input
            validate_args(jsonData, **kwargs)
        except Exception as err:
            errorReturn = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (
                SystemExit(err), kwargs['ws'], kwargs['row_num'])
            raise ErrException(errorReturn)

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        kwargs["multi_select"] = False
        jsonVars = kwargs['easy_jsonData']['components']['schemas']['easy_aci']['allOf'][1]['properties']
        # Prompt User for the Version of the Controller
        if templateVars['controller_type'] == 'apic':
            # APIC Version
            kwargs["var_description"] = f"Select the Version that Most Closely matches "\
                f'your version for the Site "{templateVars["site_name"]}".'
            kwargs["jsonVars"] = jsonVars['apic_versions']['enum']
            kwargs["defaultVar"] = jsonVars['apic_versions']['default']
            kwargs["varType"] = 'APIC Version'
            templateVars['version'] = variablesFromAPI(**kwargs)
        else:
            # NDO Version
            kwargs["var_description"] = f'Select the Version that Most Closely matches '\
                f'your version for the Site "{templateVars["site_name"]}".'
            kwargs["jsonVars"] = jsonVars['easyDict']['latest_versions']['ndo_versions']['enum']
            kwargs["defaultVar"] = jsonVars['easyDict']['latest_versions']['ndo_versions']['default']
            kwargs["varType"] = 'NDO Version'
            templateVars['version'] = variablesFromAPI(**kwargs)

        # Save the Site Information into Environment Variables
        site_id = 'site_id_%s' % (kwargs['site_id'])
        os.environ[site_id] = '%s' % (templateVars)

        # # If the state_location is tfc configure workspaces in the cloud
        # if kwargs['run_location'] == 'tfc' and kwargs['configure_terraform_cloud'] == 'true':
        #     # Initialize the Class
        #     class_init = '%s()' % ('lib_terraform.Terraform_Cloud')
        # 
        #     # Get terraform_cloud_token
        #     terraform_cloud().terraform_token()
        # 
        #     # Get workspace_ids
        #     easy_jsonData = kwargs['easy_jsonData']
        #     terraform_cloud().create_terraform_workspaces(easy_jsonData, folder_list, kwargs["site_name"])
        # 
        #     if kwargs['auth_type'] == 'user_pass' and kwargs["controller_type"] == 'apic':
        #         var_list = ['apicUrl', 'aciUser', 'aciPass']
        #     elif kwargs["controller_type"] == 'apic':
        #         var_list = ['apicUrl', 'certName', 'privateKey']
        #     else:
        #         var_list = ['ndoUrl', 'ndoDomain', 'ndoUser', 'ndoPass']
        # 
        #     # Get var_ids
        #     tf_var_dict = {}
        #     for folder in folder_list:
        #         folder_id = 'site_id_%s_%s' % (kwargs['site_id'], folder)
        #         # kwargs['workspace_id'] = workspace_dict[folder_id]
        #         kwargs['description'] = ''
        #         # for var in var_list:
        #         #     tf_var_dict = tf_variables(class_init, folder, var, tf_var_dict, **kwargs)
        # 
        # site_wb = '%s_intf_selectors.xlsx' % (kwargs['site_name'])
        # if not os.path.isfile(site_wb):
        #     kwargs['wb'].save(filename=site_wb)
        #     wb_wr = load_workbook(site_wb)
        #     ws_wr = wb_wr.get_sheet_names()
        #     for sheetName in ws_wr:
        #         if sheetName not in ['Sites']:
        #             sheetToDelete = wb_wr.get_sheet_by_name(sheetName)
        #             wb_wr.remove_sheet(sheetToDelete)
        #     wb_wr.save(filename=site_wb)

        # Return Dictionary
        kwargs['easyDict'] = OrderedDict(sorted(kwargs['easyDict'].items()))
        return kwargs['easyDict']

    #======================================================
    # Function - Site Groups
    #======================================================
    def group_id(self, **kwargs):
        # Get Variables from Library
        jsonData = kwargs['easy_jsonData']['components']['schemas']['site.Groups']['allOf'][1]['properties']

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(jsonData['required_args'], jsonData['optional_args'], **kwargs)

        for x in range(1, 16):
            site = 'site_%s' % (x)
            if not kwargs[site] == None:
                validating.site_group('site_group', **kwargs)

        grp_count = 0
        for x in list(map(chr, range(ord('A'), ord('F')+1))):
            grp = 'Grp_%s' % (x)
            if kwargs['site_group'] == grp:
                grp_count += 1
        if grp_count == 0:
            ws = kwargs['ws']
            row_num = kwargs['row_num']
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Worksheet {ws.title}, Row {row_num} group, group_name "{kwargs["group"]}" is invalid.')
            print(f'   A valid Group Name is Grp_A thru Grp_F.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()

        # Save the Site Information into Environment Variables
        group_id = '%s' % (kwargs['site_group'])
        os.environ[group_id] = '%s' % (templateVars)

        # Return Dictionary
        kwargs['easyDict'] = OrderedDict(sorted(kwargs['easyDict'].items()))
        return kwargs['easyDict']