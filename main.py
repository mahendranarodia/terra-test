#!/usr/bin/env python3
"""ACI/NDO IAC - 
This Script is to Create Terraform HCL configuration from an Excel Spreadsheet.
It uses argparse to take in the following CLI arguments:
    d or dir:           Base Directory to use for creation of the HCL Configuration Files
    w or workbook:   Name of Excel Workbook file for the Data Source
"""

import argparse
import json
import lib_aci
import os
import re
import sys
import stdiomask
import subprocess
import time
from class_system_settings import system_settings, site_policies
from easy_functions import countKeys, findKeys, findVars, stdout_log
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from git import Repo

# Global Variables
excel_workbook = None
home = Path.home()
Parser = argparse.ArgumentParser(description='Intersight Easy IMM Deployment Module')
workspace_dict = {}

access_regex = re.compile('(^aep_profile|bpdu|cdp|(fibre|port)_(channel|security)|l2_interface|l3_domain|(leaf|spine)_pg|link_level|lldp|mcp|pg_(access|breakout|bundle|spine)|phys_dom|stp|vlan_pool$)')
admin_regex = re.compile('(^export_policy|firmware|maint_group|radius|realm|remote_host|security|tacacs|tacacs_acct$)')
system_settings_regex = re.compile('(^apic_preference|bgp_(asn|rr)|global_aes$)')
bridge_domains_regex = re.compile('(^add_bd$)')
contracts_regex = re.compile('(^(contract|filter|subject)_(add|entry|to_epg)$)')
dhcp_regex = re.compile('(^dhcp_add$)')
epgs_regex = re.compile('(^add_epg$)')
fabric_regex = re.compile('(^date_time|dns|dns_profile|domain|pod_policy|ntp|sch_dstgrp|sch_receiver|snmp_(client|clgrp|comm|policy|trap|user)|syslog_(dg|rmt)|trap_groups$)')
inventory_regex = re.compile('(^apic_inb|switch|vpc_pair$)')
l3out_regex = re.compile('(^add_l3out|ext_epg|node_(prof|intf|path)|bgp_peer$)')
mgmt_tenant_regex = re.compile('(^add_bd|mgmt_epg|oob_ext_epg$)')
sites_regex = re.compile('(^site_id|group_id$)')
tenant_regex = re.compile('(^add_tenant$)')
vrfs_regex = re.compile('(^add_vrf|ctx_common$)')
vmm_regex = re.compile('(^add_vrf|ctx_common$)')

def apply_aci_terraform(folders):

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Found the Followng Folders with uncommitted changes:\n')
    for folder in folders:
        print(f'  - {folder}')
    print(f'\n  Beginning Terraform Plan and Apply in each folder.')
    print(f'\n-----------------------------------------------------------------------------\n')

    time.sleep(7)

    response_p = ''
    response_a = ''
    for folder in folders:
        path = './%s' % (folder)
        lock_count = 0
        p = subprocess.Popen(['terraform', 'init', '-plugin-dir=../../../terraform-plugins/providers/'],
                             cwd=path,
                             stdout=subprocess.PIPE,
                             stderr=subprocess.STDOUT)
        for line in iter(p.stdout.readline, b''):
            print(line)
            if re.search('does not match configured version', line.decode("utf-8")):
                lock_count =+ 1

        if lock_count > 0:
            p = subprocess.Popen(['terraform', 'init', '-upgrade', '-plugin-dir=../../../terraform-plugins/providers/'], cwd=path)
            p.wait()
        p = subprocess.Popen(['terraform', 'plan', '-out=main.plan'], cwd=path)
        p.wait()
        while True:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'  Terraform Plan Complete.  Please Review the Plan and confirm if you want')
            print(f'  to move forward.  "A" to Apply the Plan. "S" to Skip.  "Q" to Quit.')
            print(f'  Current Working Directory: {folder}')
            print(f'\n-----------------------------------------------------------------------------\n')
            response_p = input('  Please Enter ["A", "S" or "Q"]: ')
            if re.search('^(A|S)$', response_p):
                break
            elif response_p == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "A", "S" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

        if response_p == 'A':
            p = subprocess.Popen(['terraform', 'apply', '-parallelism=1', 'main.plan'], cwd=path)
            p.wait()

        while True:
            if response_p == 'A':
                response_p = ''
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Terraform Apply Complete.  Please Review for any errors and confirm if you')
                print(f'  want to move forward.  "M" to Move to the Next Section. "Q" to Quit..')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_a = input('  Please Enter ["M" or "Q"]: ')
            elif response_p == 'S':
                break
            if response_a == 'M':
                break
            elif response_a == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "M" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

def check_git_status():
    random_folders = []
    git_path = './'
    result = subprocess.Popen(['python3', '-m', 'git_status_checker'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while(True):
        # returns None while subprocess is running
        retcode = result.poll()
        line = result.stdout.readline()
        line = line.decode('utf-8')
        if re.search(r'M (.*/).*.tf\n', line):
            folder = re.search(r'M (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (.*/).*.tf\n', line):
            folder = re.search(r'\?\? (.*/).*.tf\n', line).group(1)
            if not re.search(r'ACI.templates', folder):
                if not folder in random_folders:
                    random_folders.append(folder)
        elif re.search(r'\?\? (ACI/.*/)\n', line):
            folder = re.search(r'\?\? (ACI/.*/)\n', line).group(1)
            if not (re.search(r'ACI.templates', folder) or re.search(r'\.terraform', folder)):
                if os.path.isdir(folder):
                    folder = [folder]
                    random_folders = random_folders + folder
                else:
                    group_x = [os.path.join(folder, o) for o in os.listdir(folder) if os.path.isdir(os.path.join(folder,o))]
                    random_folders = random_folders + group_x
        if retcode is not None:
            break

    if not random_folders:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   There were no uncommitted changes in the environment.')
        print(f'   Proceedures Complete!!! Closing Environment and Exiting Script.')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()

    strict_folders = []
    folder_order = ['Access', 'System', 'Tenant_common', 'Tenant_infra', 'Tenant_mgmt', 'Fabric', 'Admin', 'VLANs', 'Tenant_infra',]
    for folder in folder_order:
        for fx in random_folders:
            if folder in fx:
                if 'ACI' in folder:
                    strict_folders.append(fx)
    for folder in strict_folders:
        if folder in random_folders:
            random_folders.remove(folder)
    for folder in random_folders:
        if 'ACI' in folder:
            strict_folders.append(folder)

    # print(strict_folders)
    return strict_folders

def get_user_pass():
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   Beginning Proceedures to Apply Terraform Resources to the environment')
    print(f'\n-----------------------------------------------------------------------------\n')

    user = input('Enter APIC username: ')
    while True:
        try:
            password = stdiomask.getpass(prompt='Enter APIC password: ')
            break
        except Exception as e:
            print('Something went wrong. Error received: {}'.format(e))

    os.environ['TF_VAR_aciUser'] = '%s' % (user)
    os.environ['TF_VAR_aciPass'] = '%s' % (password)

def process_access(easyDict, easy_jsonData, wb):
    # Evaluate Access Worksheet
    lib_aci_ref = 'lib_aci.Access_Policies'
    func_regex = access_regex
    ws = wb['Access']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_admin(easyDict, easy_jsonData, wb):
    # Evaluate Admin Worksheet
    lib_aci_ref = 'lib_aci.Admin_Policies'
    func_regex = admin_regex
    ws = wb['Admin']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_bridge_domains(easyDict, easy_jsonData, wb):
    # Evaluate Bridge_Domains Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = bridge_domains_regex
    ws = wb['Bridge_Domains']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_contracts(easyDict, easy_jsonData, wb):
    # Evaluate Contracts Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = contracts_regex
    ws = wb['Contracts']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_dhcp_relay(easyDict, easy_jsonData, wb):
    # Evaluate DHCP Relay Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = dhcp_regex
    ws = wb['DHCP Relay']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_epgs(easyDict, easy_jsonData, wb):
    # Evaluate EPGs Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = epgs_regex
    ws = wb['EPGs']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_fabric(easyDict, easy_jsonData, wb):
    # Evaluate Fabric Worksheet
    class_init = 'fabric'
    class_folder = 'fabric'
    func_regex = fabric_regex
    ws = wb['Fabric']
    read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)

def process_inventory(easyDict, easy_jsonData, wb):
    # Evaluate Inventory Worksheet
    lib_aci_ref = 'lib_aci.Access_Policies'
    func_regex = inventory_regex
    ws = wb['Inventory']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_l3out(easyDict, easy_jsonData, wb):
    # Evaluate L3Out Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = l3out_regex
    ws = wb['L3Out']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_mgmt_tenant(easyDict, easy_jsonData, wb):
    # Evaluate Mgmt_Tenant Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    ws = wb['Mgmt_Tenant']
    func_regex = mgmt_tenant_regex
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_sites(easyDict, easy_jsonData, wb):
    # Evaluate Sites Worksheet
    class_init = 'site_policies'
    class_folder = 'sites'
    func_regex = sites_regex
    ws = wb['Sites']
    read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)

def process_system_settings(easyDict, easy_jsonData, wb):
    # Evaluate System_Settings Worksheet
    class_init = 'system_settings'
    class_folder = 'system_settings'
    func_regex = system_settings_regex
    ws = wb['System_Settings']
    read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws)

def process_tenants(easyDict, easy_jsonData, wb):
    # Evaluate Tenants Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = tenant_regex
    ws = wb['Tenants']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_vrfs(easyDict, easy_jsonData, wb):
    # Evaluate VRF Worksheet
    lib_aci_ref = 'lib_aci.Tenant_Policies'
    func_regex = vrfs_regex
    ws = wb['VRF']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def process_vmm(easyDict, easy_jsonData, wb):
    # Evaluate Sites Worksheet
    lib_aci_ref = 'lib_aci.VMM_Policies'
    func_regex = vmm_regex
    ws = wb['VMM']
    read_worksheet(wb, ws, lib_aci_ref, func_regex)

def read_worksheet(class_init, class_folder, easyDict, easy_jsonData, func_regex, wb, ws):
    class_folder = class_folder
    rows = ws.max_row
    func_list = findKeys(ws, func_regex)
    stdout_log(ws, None)
    for func in func_list:
        count = countKeys(ws, func)
        var_dict = findVars(ws, func, rows, count)
        for pos in var_dict:
            row_num = var_dict[pos]['row']
            del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            stdout_log(ws, row_num)
            var_dict[pos].update(
                {
                    'easyDict':easyDict,
                    'easy_jsonData':easy_jsonData,
                    'row_num':row_num,
                    'wb':wb,
                    'ws':ws
                }
            )
            easyDict = eval(f"{class_init}(class_folder).{func}(**var_dict[pos])")

def wb_update(wr_ws, status, i):
    # build green and red style sheets for excel
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wsh1 = NamedStyle(name="wsh1")
    wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wsh2 = NamedStyle(name="wsh2")
    wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wsh2.fill = PatternFill("solid", fgColor="305496")
    wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    green_st = NamedStyle(name="ws_odd")
    green_st.alignment = Alignment(horizontal="center", vertical="center")
    green_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    green_st.fill = PatternFill("solid", fgColor="D9E1F2")
    green_st.font = Font(bold=False, size=12, color="44546A")
    red_st = NamedStyle(name="ws_even")
    red_st.alignment = Alignment(horizontal="center", vertical="center")
    red_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    red_st.font = Font(bold=False, size=12, color="44546A")
    yellow_st = NamedStyle(name="ws_even")
    yellow_st.alignment = Alignment(horizontal="center", vertical="center")
    yellow_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    yellow_st.font = Font(bold=False, size=12, color="44546A")
    # green_st = xlwt.easyxf('pattern: pattern solid;')
    # green_st.pattern.pattern_fore_colour = 3
    # red_st = xlwt.easyxf('pattern: pattern solid;')
    # red_st.pattern.pattern_fore_colour = 2
    # yellow_st = xlwt.easyxf('pattern: pattern solid;')
    # yellow_st.pattern.pattern_fore_colour = 5
    # if stanzas to catch the status code from the request
    # and then input the appropriate information in the workbook
    # this then writes the changes to the doc
    if status == 200:
        wr_ws.write(i, 1, 'Success (200)', green_st)
    if status == 400:
        print("Error 400 - Bad Request - ABORT!")
        print("Probably have a bad URL or payload")
        wr_ws.write(i, 1, 'Bad Request (400)', red_st)
        pass
    if status == 401:
        print("Error 401 - Unauthorized - ABORT!")
        print("Probably have incorrect credentials")
        wr_ws.write(i, 1, 'Unauthorized (401)', red_st)
        pass
    if status == 403:
        print("Error 403 - Forbidden - ABORT!")
        print("Server refuses to handle your request")
        wr_ws.write(i, 1, 'Forbidden (403)', red_st)
        pass
    if status == 404:
        print("Error 404 - Not Found - ABORT!")
        print("Seems like you're trying to POST to a page that doesn't"
              " exist.")
        wr_ws.write(i, 1, 'Not Found (400)', red_st)
        pass
    if status == 666:
        print("Error - Something failed!")
        print("The POST failed, see stdout for the exception.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass
    if status == 667:
        print("Error - Invalid Input!")
        print("Invalid integer or other input.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass

def main():
    description = None
    if description is not None:
        Parser.description = description
    Parser.add_argument('-d', '--dir', default='ACI',
                        help='The Directory to Publish the Terraform Files to.'
    )
    Parser.add_argument('-wb', '--workbook', default='ACI_Base_Workbookv2.xlsx',
                        help='The Workbook to read for Input.'
    )
    Parser.add_argument('-ws', '--worksheet', default=None,
                        help='The Workbook to read for Input.'
    )
    args = Parser.parse_args()

    jsonFile = 'templates/variables/easy_variables.json'
    jsonOpen = open(jsonFile, 'r')
    easy_jsonData = json.load(jsonOpen)
    jsonOpen.close()

    # Ask user for required Information: ACI_DEPLOY_FILE
    if os.path.isfile(args.workbook):
        excel_workbook = args.workbook
    else:
        print('\nWorkbook not Found.  Please enter a valid /path/filename for the source workbook you will be using.')
        while True:
            print('Please enter a valid /path/filename for the source you will be using.')
            excel_workbook = input('/Path/Filename: ')
            if os.path.isfile(excel_workbook):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')

    # Load Workbook
    wb = lib_aci.read_in(excel_workbook)

    # Create Dictionary for Worksheets in the Workbook

    easyDict = {
        'access':{},
        'admin':{},
        'fabric':{
            'date_and_time':[],
            'dns_profiles':[],
            'date_and_time':[],
            'date_and_time':[],
        },
        'inventory':{},
        'sites':{},
        'tenants':{},
        'system_settings':{}
    }

    # Run Proceedures for Worksheets in the Workbook
    process_sites(easyDict, easy_jsonData, wb)

    # Either Run All Remaining Proceedures or Just Specific based on sys.argv[2:]
    if not args.worksheet == None:
        if re.search('site', str(args.worksheet)):
            process_sites(easyDict, easy_jsonData, wb)
        elif re.search('access', str(args.worksheet)):
            process_access(easyDict, easy_jsonData, wb)
        elif re.search('admin', str(args.worksheet)):
            process_admin(easyDict, easy_jsonData, wb)
        elif re.search('inventory', str(args.worksheet)):
            process_inventory(easyDict, easy_jsonData, wb)
        elif re.search('system_settings', str(args.worksheet)):
            process_system_settings(easyDict, easy_jsonData, wb)
        elif re.search('fabric', str(args.worksheet)):
            process_fabric(easyDict, easy_jsonData, wb)
        elif re.search('tenant', str(args.worksheet)):
            process_tenants(easyDict, easy_jsonData, wb)
        elif re.search('vrf', str(args.worksheet)):
            process_vrfs(easyDict, easy_jsonData, wb)
        elif re.search('contract', str(args.worksheet)):
            process_contracts(easyDict, easy_jsonData, wb)
        elif re.search('l3out', str(args.worksheet)):
            process_l3out(easyDict, easy_jsonData, wb)
        elif re.search('mgmt', str(args.worksheet)):
            process_mgmt_tenant(easyDict, easy_jsonData, wb)
        elif re.search('bd', str(args.worksheet)):
            process_bridge_domains(easyDict, easy_jsonData, wb)
        elif re.search('dhcp', str(args.worksheet)):
            process_dhcp_relay(easyDict, easy_jsonData, wb)
        elif re.search('epg', str(args.worksheet)):
            process_epgs(easyDict, easy_jsonData, wb)
        elif re.search('vmm', str(args.worksheet)):
            process_vmm(easyDict, easy_jsonData, wb)
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   {args.worksheet} is not a valid worksheet.  If you are trying to run')
            print(f'   a single worksheet please re-enter the -ws argument.  Exiting...')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        process_system_settings(easyDict, easy_jsonData, wb)
        exit()
        process_fabric(easyDict, easy_jsonData, wb)
        process_admin(easyDict, easy_jsonData, wb)
        process_access(easyDict, easy_jsonData, wb)
        process_inventory(easyDict, easy_jsonData, wb)
        process_tenants(easyDict, easy_jsonData, wb)
        process_l3out(easyDict, easy_jsonData, wb)
        process_contracts(easyDict, easy_jsonData, wb)
        process_mgmt_tenant(easyDict, easy_jsonData, wb)
        process_vrfs(easyDict, easy_jsonData, wb)
        process_bridge_domains(easyDict, easy_jsonData, wb)
        # process_dhcp_relay(easyDict, easy_jsonData, wb)
        process_epgs(easyDict, easy_jsonData, wb)
        # process_vmm(easyDict, easy_jsonData, wb)

    folders = check_git_status()
    get_user_pass()
    apply_aci_terraform(folders)
    # else:
    #     print('hello')
    #     path = './'
    #     repo = Repo.init(path)

    #     index = Repo.init(path.index)

    #     index.commit('Testing Commit')

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Proceedures Complete!!! Closing Environment and Exiting Script.')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()

if __name__ == '__main__':
    main()
